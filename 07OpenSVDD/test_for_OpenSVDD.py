import argparse
import importlib
import warnings
import glob
import os
import joblib
import pandas as pd
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import f1_score, roc_auc_score, roc_curve
from sklearn.preprocessing import OneHotEncoder
from torchvision import transforms
import torch.optim as optim
from torch.utils.data import dataloader
import random
from torchvision import transforms, utils
from torch.utils.data import Dataset, DataLoader
from PIL import Image
import os
import numpy as np
from matplotlib import pyplot as plt, pyplot
import sys
import time
import shutil
import errno

import numpy as np
# import pandas as pd

from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn.metrics import cohen_kappa_score


from dataset.dataset2 import dataset_for_image, dataset_for_radar, dataset_for_ADSB
from models.resnet50_for_1d import ResNet50
# from utils.openSVDD import openSVDD

parser = argparse.ArgumentParser("Training")

# misc
parser.add_argument('--eval_freq', type=int, default=1)
parser.add_argument('--print-freq', type=int, default=100)
parser.add_argument('--gpu', type=str, default='0')
parser.add_argument('--seed', type=int, default=0)

warnings.filterwarnings('ignore')


def seed_everything(TORCH_SEED):
    random.seed(TORCH_SEED)
    os.environ['PYTHONHASHSEED'] = str(TORCH_SEED)
    np.random.seed(TORCH_SEED)
    torch.manual_seed(TORCH_SEED)
    torch.cuda.manual_seed_all(TORCH_SEED)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def test(net, testloader, opensvdd, **options):
    net.eval()
    correct, total = 0, 0
    correct_bin = 0
    torch.cuda.empty_cache()
    label_list, outputs_list, label_binary = [], [], []
    with torch.no_grad():
        for data, target_name, target in testloader:
            if options['use_gpu']:
                data = data.cuda()
            with torch.set_grad_enabled(False):
                y = net(data)
                outputs_list.append(y.cpu().numpy())
                label_list.extend(target)
    outputs_list = np.concatenate(outputs_list, 0)
    label_list = np.array(label_list)

    labels_binary = [0 if x >= len(options['train_typelist']) else 1 for x in label_list]

    label_list[label_list >= len(options['train_typelist'])] = -1
    labels_open_pred, scores, _, norm_distance = opensvdd.predict(outputs_list, label_list, options['threshold'])
    labels_open_binpred = [1 if x >= 0 else 0 for x in labels_open_pred]
    labels_binary = np.array(labels_binary)
    labels_open_binpred = np.array(labels_open_binpred)
    scores = 1. / scores.A
    evaluate = classification_report(label_list, labels_open_pred, output_dict=True)
    accuracy = evaluate['accuracy']
    f1_weighted = evaluate['weighted avg']['f1-score']
    f1_macro = evaluate['macro avg']['f1-score']
    AUROC_bin = roc_auc_score(labels_binary, scores)
    # AUROC_bin = 0

    print(f"OpenSVDD AUROC is %.3f" % (100*AUROC_bin))
    print(f"OpenSVDD accuracy is %.3f" % (100*accuracy))
    print(f"OpenSVDD Weighted F1 is %.3f" % (100*f1_weighted))
    print(f"OpenSVDD Macro F1 is %.3f" % (100*f1_macro))
    print(f"_________________________________________")

    return 100*AUROC_bin, 100*accuracy, 100*f1_weighted, 100*f1_macro


def get_threshold(model, opensvdd, val_loader, **options):
    model.eval()
    classwise_thresholds = []
    classwise_logits = []
    scores, outputs_list, label_list = [], [], []
    torch.cuda.empty_cache()
    for i in range(len(options['train_typelist'])):
        classwise_logits.append([])
    with torch.no_grad():
        for data in val_loader:
            inputs, _, target = data
            inputs, target = inputs.cuda(), target.numpy()
            inputs = torch.squeeze(inputs, dim=1)
            with torch.set_grad_enabled(False):
                y = model(inputs)
                outputs_list.append(y.cpu().numpy())
                label_list.extend(target)
        outputs_list = np.concatenate(outputs_list, 0)
        label_list = np.array(label_list)
    labels_open_pred, scores, _, norm_distance = opensvdd.predict(outputs_list, label_list)
    norm_distance = np.array(norm_distance)
    minIndexes = np.argmin(norm_distance, axis=1)
    for j, label in enumerate(label_list):
        if minIndexes[j] == label:
            classwise_logits[label].append(norm_distance[j, label].item())
    for val in classwise_logits:
        if len(val) == 0:
            classwise_thresholds.append(0)
        else:
            threshold = np.percentile(val, 98)
            classwise_thresholds.append(threshold)
    print(classwise_thresholds)
    return classwise_thresholds


def main(options):
    seed_everything(options['seed'])
    options['num_classes'] = len(options['train_typelist'])
    feat_dim = len(options['train_typelist'])
    # Model
    print('==> Building model: '+options['model_name'])
    if options['dataset'] == 'radar':
        model = ResNet50(in_channels=2, feature_dim=len(options['train_typelist']), classes=len(options['train_typelist']))
    else:
        model = ResNet50(in_channels=3, feature_dim=len(options['train_typelist']), classes=len(options['train_typelist']))
    # print(model)

    options.update(
        {
            'use_gpu': True,
            'feat_dim': feat_dim,
        }
    )

    Loss = importlib.import_module('loss.' + options['loss'])
    criterion = getattr(Loss, options['loss'])(**options)
    model = nn.DataParallel(model).cuda()
    criterion = criterion.cuda()
    params = [{'params': model.parameters()},
              {'params': criterion.parameters()}]
    print(options['model_name'])
    loss_weight_path = options['loss_path']
    criterion.load_state_dict(torch.load(loss_weight_path))
    backbone_weight_path = options['weight_path']
    model.load_state_dict(torch.load(backbone_weight_path))
    opensvdd = joblib.load(options['SVDD_path'])

    data_root_path = options['data_root'] + options['dataset']

    # Dataset
    if options['dataset'] == 'radar':
        valset = dataset_for_radar(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_radar(data_root_path, options['test_typelist'], mode_type='test')
    else:
        valset = dataset_for_ADSB(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_ADSB(data_root_path, options['test_typelist'], mode_type='val')

    #  dataloader
    valloader = DataLoader(valset, batch_size=options['batch_size'], shuffle=False, pin_memory=True)
    testloader = DataLoader(testset, batch_size=options['batch_size'], shuffle=False, pin_memory=True)
    print("All test data:", len(testset))

    threshold = get_threshold(model, opensvdd, valloader, **options)
    options.update(
        {
            'threshold': threshold  # threshold_list[i]
        })
    AUROC_bin, accuracy, f1_weighted, f1_macro = test(model, testloader, opensvdd, **options)
    new_test_results = {
        'AUROC': [AUROC_bin],
        'Accuracy': [accuracy],
        'F1_weighted': [f1_weighted],
        'F1_macro': [f1_macro],
    }
    new_df = pd.DataFrame(new_test_results)
    output_file = 'TEST.xlsx'
    try:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            startrow = writer.sheets['Sheet1'].max_row
            new_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)
        print(f"New test results have been appended to {output_file}")
    except FileNotFoundError:
        new_df.to_excel(output_file, index=False)
        print(f"The file does not exist. A new file has been created: {output_file}")

    return 0

if __name__ == '__main__':
    args = parser.parse_args()
    options = vars(args)
    results = dict()

    from split import train_splits as train_splits
    from split import test_splits as test_splits

    model_name = 'ResNet50'
    test_list = ['radar', 'ADSB']
    for j in range(len(test_list)):
        torch.cuda.empty_cache()
        options.update(
            {
                'dataset': test_list[j]
            }
        )

        result_model_path = './log/ARPL/' + model_name + '/' + options['dataset']
        files = os.listdir(result_model_path)

        for i in range(len(train_splits[options['dataset']])):
            train_split = train_splits[options['dataset']][i]
            test_split = test_splits[options['dataset']][i]

            weight_path = './log/ARPL/' + model_name + '/' + options['dataset'] + '/' + files[i] + '/result_model/models/checkpoints/ResNet1D_ARPLoss_.pth'
            loss_path = './log/ARPL/' + model_name + '/' + options['dataset'] + '/' + files[i] + '/result_model/models/checkpoints/ResNet1D_ARPLoss__criterion.pth'
            SVDD_path = './log/ARPL/' + model_name + '/' + options['dataset'] + '/' + files[i] + '/result_model/models/checkpoints/opensvdd3.pkl'

            options.update(
                {
                    'train_typelist': train_split,  # [0,1,2,3,4,5,6]
                    'val_typelist': train_split,
                    'test_typelist': test_split,
                    'seed': 0,
                    'batch_size': 64,
                    'weight_path': weight_path,
                    'loss_path': loss_path,
                    'data_root': 'D:/wu/00data/',
                    'SVDD_path': SVDD_path,
                }
            )
            main(options)