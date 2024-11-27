import argparse
import importlib
import warnings
import glob
import os
import joblib
import pandas as pd
import torch
import torch.nn as nn
from openpyxl import load_workbook
import random
from torchvision import transforms, utils
from torch.utils.data import Dataset, DataLoader
from PIL import Image
import os
import numpy as np
from matplotlib import pyplot as plt
import sys
import time
import shutil
import errno
import numpy as np
# import pandas as pd
from sklearn.metrics import confusion_matrix, roc_auc_score
from sklearn.metrics import classification_report
from sklearn.metrics import cohen_kappa_score
from dataset.dataset2 import dataset_for_image, dataset_for_radar, dataset_for_ADSB
from models.resnet50_for_1d import ResNet50

parser = argparse.ArgumentParser("Training")

# Dataset
parser.add_argument('--known_typelist', default=[0, 1, 2, 3, 4, 5])  # [0, 1, 2, 3, 4, 5, 6, 7]
# parser.add_argument('--out-num', type=int, default=50, help='For CIFAR100')

# optimization
parser.add_argument('--batch_size', type=int, default=64)
parser.add_argument('--lr', type=float, default=0.0001, help="learning rate for model")
parser.add_argument('--maxepoch', type=int, default=10)
parser.add_argument('--best_auroc_val', type=float, default=0)
parser.add_argument('--loss', type=str, default='ARPLoss')
# parser.add_argument('--threshold', default=0.8)
parser.add_argument('--temp', type=float, default=1.0, help="temp")

# model
parser.add_argument('--weight-pl', type=float, default=0.1, help="weight for center loss")
parser.add_argument('--beta', type=float, default=0.1, help="weight for entropy loss")
parser.add_argument('--model_name', type=str, default='ResNet50')
parser.add_argument('--weight_path', type=str, default='./log/ResNet50_loss1.13474.pth')
# parser.add_argument('--feature_dim', default=256)

# misc
parser.add_argument('--eval_freq', type=int, default=1)
parser.add_argument('--print-freq', type=int, default=100)
parser.add_argument('--gpu', type=str, default='0')
parser.add_argument('--seed', type=int, default=0)

warnings.filterwarnings('ignore')


def seed_everything(TORCH_SEED):
    random.seed(TORCH_SEED)
    os.environ['PYTHONHASHSEED'] = str(TORCH_SEED)
    np.random.seed(TORCH_SEED)
    torch.manual_seed(TORCH_SEED)
    torch.cuda.manual_seed_all(TORCH_SEED)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def test(net, criterion, testloader, threshold, epoch=None, **options):
    net.eval()
    correct, total = 0, 0

    torch.cuda.empty_cache()

    _pred_k, _pred_u, _labels = [], [], []
    scores, labels, predicted, max_value_list, max_value_idx_list, outputs_list = [], [], [], [], [], []

    with torch.no_grad():
        for data, target_name, target in testloader:
            if options['use_gpu']:
                data, target = data.cuda(), target.cuda()

            with torch.set_grad_enabled(False):  # 与上同理
                # ------------------------------------------------------------
                y = net(data)
                logits, _ = criterion(x=y, y=None)
                logits = logits.cpu().numpy()
                max_value = np.max(logits, axis=1)
                max_value_idx = np.argmax(logits, axis=1)
                labels.append(target)
                outputs_list.append(logits)
                predicted.extend(max_value_idx)
                max_value_list.extend(max_value)
                max_value_idx_list.extend(max_value_idx)
                scores.extend(np.max(logits, axis=1))
                total += target.size(0)
    labels = torch.cat(labels, dim=0).cpu().numpy()
    labels = np.array(labels)
    predicted = np.array(predicted)
    max_value_list = np.array(max_value_list)
    outputs_list = np.concatenate(outputs_list, 0)
    scores = np.array(scores)
    # np.savez('./log/ARPL_feature.npz', outputs_list=outputs_list, label_list=labels)

    labels_open = []
    for i in range(0, len(labels)):
        if labels[i] >= len(options['train_typelist']):
            labels[i] = len(options['train_typelist'])
            labels_open.append(0)
        else:
            labels_open.append(1)

    labels_open = np.array(labels_open)

    unknown_index = np.where(max_value_list < threshold)[0]
    labels_open_pred = predicted
    for i in range(len(unknown_index)):
        labels_open_pred[unknown_index[i]] = len(options['train_typelist'])

    evaluate = classification_report(labels, labels_open_pred, output_dict=True)
    accuracy = evaluate['accuracy']
    f1_weighted = evaluate['weighted avg']['f1-score']
    f1_macro = evaluate['macro avg']['f1-score']
    AUROC_bin = roc_auc_score(labels_open, scores)
    # AUROC_bin = 0

    print(f"ARPL AUROC is %.3f" % (100*AUROC_bin))
    print(f"ARPL accuracy is %.3f" % (100*accuracy))
    print(f"ARPL Weighted F1 is %.3f" % (100*f1_weighted))
    print(f"ARPL Macro F1 is %.3f" % (100*f1_macro))
    print(f"_________________________________________")

    return 100*AUROC_bin,100*accuracy,100*f1_weighted,100*f1_macro

def get_threshold(model, criterion, val_loader, **options):
    model.eval()
    classwise_logits = []
    torch.cuda.empty_cache()
    for i in range(len(options['train_typelist'])):
        classwise_logits.append([])
    with torch.no_grad():
        for data in val_loader:
            inputs, _, target = data
            inputs, target = inputs.cuda(), target.cuda()
            inputs = torch.squeeze(inputs, dim=1)
            with torch.set_grad_enabled(False):
                y = model(inputs)
                logits, _ = criterion(y, None)
                maxLogit, maxIndexes = torch.max(logits, 1)

                for j, label in enumerate(target):
                    if maxIndexes[j] == label:
                        classwise_logits[label].append(logits[j, label].item())
    classwise_logits = np.concatenate(classwise_logits)
    sorted_values = np.sort(classwise_logits)
    threshold = sorted_values[int(0.03 * len(classwise_logits))]
    threshold = np.array(threshold).squeeze()
    print(threshold)
    return threshold


def main(options):
    device = torch.device('cuda:0')

    seed_everything(options['seed'])
    options['num_classes'] = len(options['train_typelist'])
    feat_dim = len(options['train_typelist'])
    # Model
    print('==> Building model: '+options['model_name'])
    if options['dataset'] == 'radar':
        net = ResNet50(in_channels=2, num_classes=len(options['train_typelist']))
    elif options['dataset'] == 'ADSB':
        net = ResNet50(in_channels=3, num_classes=len(options['train_typelist']))

    options.update(
        {
            'use_gpu': True,
            'feat_dim': feat_dim,
        }
    )

    Loss = importlib.import_module('loss.' + options['loss'])
    criterion = getattr(Loss, options['loss'])(**options)
    net = nn.DataParallel(net).cuda()
    criterion = criterion.cuda()
    params = [{'params': net.parameters()},
              {'params': criterion.parameters()}]
    print(options['model_name'])
    loss_weight_path = options['loss_path']
    criterion.load_state_dict(torch.load(loss_weight_path))
    backbone_weight_path = options['weight_path']
    net.load_state_dict(torch.load(backbone_weight_path))
    data_root_path = options['data_root'] + options['dataset']

    # Dataset
    if options['dataset'] == 'radar':
        valset = dataset_for_radar(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_radar(data_root_path, options['test_typelist'], mode_type='test')
    else:
        valset = dataset_for_ADSB(data_root_path,  options['train_typelist'], mode_type='val')
        testset = dataset_for_ADSB(data_root_path, options['test_typelist'], mode_type='test')

    #  dataloader
    valloader = DataLoader(valset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All val data:", len(valset))
    testloader = DataLoader(testset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All test data:", len(testset))

    threshold = get_threshold(net, criterion, valloader, **options)
    AUROC_bin, accuracy, f1_weighted, f1_macro = test(net, criterion, testloader, threshold, epoch=None, **options)

    new_test_results = {
        'AUROC': [AUROC_bin],
        'Accuracy': [accuracy],
        'F1_weighted': [f1_weighted],
        'F1_macro': [f1_macro],
    }
    new_df = pd.DataFrame(new_test_results)
    output_file = 'TEST.xlsx'
    try:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            startrow = writer.sheets['Sheet1'].max_row
            new_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)

        print(f"New test results have been appended to {output_file}")
    except FileNotFoundError:
        new_df.to_excel(output_file, index=False)
        print(f"The file does not exist. A new file has been created: {output_file}")
    return 0

if __name__ == '__main__':
    args = parser.parse_args()
    options = vars(args)
    results = dict()
    options.update({
        'model_name': 'ResNet50'
    })
    from split import train_splits as train_splits
    from split import test_splits as test_splits
    test_list = ['radar', 'ADSB']
    for j in range(len(test_list)):
        torch.cuda.empty_cache()
        options.update(
            {
                'dataset': test_list[j]
            }
        )

        result_model_path = './log/ARPL/' + options['model_name'] + '/' + options['dataset']
        files = os.listdir(result_model_path)

        for i in range(len(train_splits[options['dataset']])):
            train_split = train_splits[options['dataset']][i]
            test_split = test_splits[options['dataset']][i]

            weight_path = './log/ARPL/' + options['model_name'] + '/' + options['dataset'] + '/' + files[i] + '/result_model/models/checkpoints/'+ options['model_name'] + '_ARPLoss_.pth'
            loss_path = './log/ARPL/' + options['model_name'] + '/' + options['dataset'] + '/' + files[i] + '/result_model/models/checkpoints/'+ options['model_name'] + '_ARPLoss__criterion.pth'

            options.update(
                {
                    'train_typelist': train_split,  # [0,1,2,3,4,5,6]
                    'val_typelist': train_split,
                    'test_typelist': test_split,
                    'seed': 0,
                    'batch_size': 256,
                    'weight_path': weight_path,
                    'loss_path': loss_path,
                    'data_root': 'D:/wu/00data/',
                }
            )

            main(options)