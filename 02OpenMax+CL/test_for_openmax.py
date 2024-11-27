import argparse
import importlib
import warnings
import glob
import os
import joblib
import pandas as pd
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import f1_score, roc_auc_score, roc_curve
from sklearn.preprocessing import OneHotEncoder
from torchvision import transforms
import torch.optim as optim
from torch.utils.data import dataloader
import random
from torchvision import transforms, utils
from torch.utils.data import Dataset, DataLoader
from PIL import Image
import os
import numpy as np
from matplotlib import pyplot as plt
import sys
import time
import shutil
import errno

import numpy as np


from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn.metrics import cohen_kappa_score


from dataset.dataset2 import dataset_for_image, dataset_for_radar, dataset_for_ADSB
from models.SKResNet import SKResNet
from utils.openmax import compute_train_score_and_mavs_and_dists,fit_weibull,openmax
from train_with_CL import seed_everything

parser = argparse.ArgumentParser("Training")

# misc
parser.add_argument('--eval_freq', type=int, default=1)
parser.add_argument('--print-freq', type=int, default=100)
parser.add_argument('--gpu', type=str, default='0')
parser.add_argument('--seed', type=int, default=0)

warnings.filterwarnings('ignore')


def test(model, device, test_loader, weibull_model, threshold, **options):
    model.eval()
    scores, labels = [], []
    with torch.no_grad():
        for data in test_loader:
            inputs, target_name, target = data
            inputs, target = inputs.to(device), target.to(device)
            outputs = model(inputs)
            scores.append(outputs)
            labels.append(target)

    scores = torch.cat(scores, dim=0).cpu().numpy()
    labels = torch.cat(labels, dim=0).cpu().numpy()
    scores = np.array(scores)[:, np.newaxis, :]
    labels = np.array(labels)

    labels_open = []
    for i in range(0, len(labels)):
        if labels[i] > len(options['train_typelist']):
            labels[i] = len(options['train_typelist'])
            labels_open.append(0)
        elif labels[i] == len(options['train_typelist']):
            labels_open.append(0)
        else:
            labels_open.append(1)

    labels_open = np.array(labels_open)
    categories = list(range(0, len(options['train_typelist'])))

    score_softmax, score_openmax = [], []
    for score in scores:
        so, ss = openmax(weibull_model, categories, score, 0.5, options['weibull_alpha'], "euclidean")  # openmax_prob, softmax_prob
        score_softmax.append(ss)
        score_openmax.append(so)
    # np.savez('./log/openmax_feature.npz', outputs_list=score_openmax, label_list=labels)
    predict = np.argmax(score_openmax, axis=1)
    max_value_list = np.max(score_openmax, axis=1)

    unknown_index = np.where(max_value_list < threshold)[0]
    labels_open_pred = predict
    for i in range(len(unknown_index)):
        labels_open_pred[unknown_index[i]] = len(options['train_typelist'])

    evaluate = classification_report(labels, labels_open_pred, output_dict=True)
    accuracy = evaluate['accuracy']
    f1_weighted = evaluate['weighted avg']['f1-score']
    f1_macro = evaluate['macro avg']['f1-score']
    AUROC_bin = roc_auc_score(labels_open, np.max(score_softmax, axis=1))

    print(f"OpenMax+CL AUROC is %.3f" % (100*AUROC_bin))
    print(f"OpenMax+CL accuracy is %.3f" % (100*accuracy))
    print(f"OpenMax+CL Weighted F1 is %.3f" % (100*f1_weighted))
    print(f"OpenMax+CL Macro F1 is %.3f" % (100*f1_macro))
    print(f"_________________________________________")

    return 100*AUROC_bin,100*accuracy,100*f1_weighted,100*f1_macro

def get_threshold(model, device, weibull_model, val_loader, **options):
    model.eval()
    classwise_logits = []
    torch.cuda.empty_cache()
    categories = list(range(0, len(options['train_typelist'])))
    for i in range(len(options['train_typelist'])):
        classwise_logits.append([])
    with torch.no_grad():
        for data in val_loader:
            inputs, _, target = data
            inputs, target = inputs.to(device), target.to(device)
            inputs = torch.squeeze(inputs, dim=1)
            logits = model(inputs)
            logits = logits.cpu()
            logits = np.array(logits)[:, np.newaxis, :]
            score_openmax = []
            for score in logits:
                so, _ = openmax(weibull_model, categories, score, 0.5, options['weibull_alpha'], "euclidean")  # openmax_prob, softmax_prob
                score_openmax.append(so)

            maxIndexes = np.argmax(score_openmax, axis=1)
            score_openmax = np.array(score_openmax)
            for j, label in enumerate(target):
                if maxIndexes[j] == label:
                    classwise_logits[label].append(score_openmax[j, label].item())
    classwise_logits = np.concatenate(classwise_logits)
    sorted_values = np.sort(classwise_logits)
    threshold = sorted_values[int(0.03 * len(classwise_logits))]
    threshold = np.array(threshold).squeeze()
    print(threshold)
    return threshold

def main(options):
    seed_everything(options['seed'])
    device = torch.device('cuda:0')
    options['num_classes'] = len(options['train_typelist'])
    feat_dim = 256
    # Model
    print('==> Building model: '+options['model_name'])
    if options['dataset'] == 'radar':
        model = SKResNet(input_channel=2, num_classes=len(options['train_typelist']))
    elif options['dataset'] == 'ADSB':
        model = SKResNet(input_channel=3, num_classes=len(options['train_typelist']))

    options.update(
        {
            'use_gpu': True,
            'feat_dim': feat_dim,
        }
    )


    model = model.to(device)
    print(options['model_name'])
    backbone_weight_path = options['weight_path']
    model.load_state_dict(torch.load(backbone_weight_path))
    data_root_path = options['data_root'] + options['dataset']
    # Dataset
    if options['dataset'] == 'radar':
        trainset = dataset_for_radar(data_root_path, options['train_typelist'], mode_type='train')
        valset = dataset_for_radar(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_radar(data_root_path, options['test_typelist'], mode_type='test')
    else:
        trainset = dataset_for_ADSB(data_root_path,  options['train_typelist'], mode_type='train')
        valset = dataset_for_ADSB(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_ADSB(data_root_path, options['test_typelist'], mode_type='test')

    #  dataloader
    trainloader = DataLoader(trainset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All training data:", len(trainset))
    validloader = DataLoader(valset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All val data:", len(trainset))
    testloader = DataLoader(testset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All test data:", len(testset))

    # Fit the weibull distribution from training data.
    print("Fittting Weibull distribution...")
    _, mavs, dists = compute_train_score_and_mavs_and_dists(len(options['train_typelist']), trainloader, device, model)
    categories = list(range(0, len(options['train_typelist'])))
    weibull_model = fit_weibull(mavs, dists, categories, options['weibull_tail'], "euclidean")
    print("Weibull distribution is fitted")

    threshold = get_threshold(model, device, weibull_model, validloader, **options)
    AUROC_bin, accuracy, f1_weighted, f1_macro = test(model, device, testloader, weibull_model, threshold, **options)

    new_test_results = {
        'AUROC': [AUROC_bin],
        'Accuracy': [accuracy],
        'F1_weighted': [f1_weighted],
        'F1_macro': [f1_macro],
    }
    new_df = pd.DataFrame(new_test_results)
    output_file = 'TEST.xlsx'
    try:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            startrow = writer.sheets['Sheet1'].max_row
            new_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)

        print(f"New test results have been appended to {output_file}")
    except FileNotFoundError:
        new_df.to_excel(output_file, index=False)
        print(f"The file does not exist. A new file has been created: {output_file}")

    return 0

if __name__ == '__main__':
    args = parser.parse_args()
    options = vars(args)
    results = dict()
    options.update({
        'model_name': 'SKResNet'
    })
    from split import train_splits as train_splits
    from split import test_splits as test_splits
    test_list = ['radar', 'ADSB']
    for j in range(len(test_list)):
        torch.cuda.empty_cache()
        options.update(
            {
                'dataset': test_list[j]
            }
        )
        result_model_path = './log/OpenMax+CL/' + options['model_name'] + '/' + options['dataset']
        files = os.listdir(result_model_path)
        for i in range(len(train_splits[options['dataset']])):
            train_split = train_splits[options['dataset']][i]
            test_split = test_splits[options['dataset']][i]

            weight_path = './log/OpenMax+CL/' + options['model_name'] + '/' + options['dataset'] + '/' + files[i] + '/result_model/'+options['model_name']+'.pth'
            options.update(
                {
                    'train_typelist': train_split,  # [0,1,2,3,4,5,6]
                    'val_typelist': train_split,
                    'test_typelist': test_split,
                    'seed': 0,
                    'batch_size': 64,
                    'weight_path': weight_path,
                    'data_root': 'D:/wu/00data/',
                    'weibull_tail': 20,
                    'weibull_alpha': len(train_split),
                }
            )

            main(options)