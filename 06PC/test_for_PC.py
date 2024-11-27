import argparse
import importlib
import warnings
import glob
import os
import joblib
import pandas as pd
import torch
import torch.nn as nn
from openpyxl import load_workbook
from sklearn.metrics import f1_score, roc_auc_score, roc_curve
from sklearn.preprocessing import OneHotEncoder
from torchvision import transforms
import torch.optim as optim
from torch.utils.data import dataloader
import random
from torchvision import transforms, utils
from torch.utils.data import Dataset, DataLoader
from PIL import Image
import os
import numpy as np
from matplotlib import pyplot as plt
import sys
import time
import shutil
import errno
import numpy as np
# import pandas as pd
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn.metrics import cohen_kappa_score
from dataset.dataset2 import dataset_for_image, dataset_for_radar, dataset_for_ADSB
from models.resnet50_for_1d import ResNet50
from train_for_PC import seed_everything


parser = argparse.ArgumentParser("Training")

# misc
parser.add_argument('--eval_freq', type=int, default=1)
parser.add_argument('--print-freq', type=int, default=100)
parser.add_argument('--gpu', type=str, default='0')
parser.add_argument('--seed', type=int, default=0)

warnings.filterwarnings('ignore')

def test(model, device, test_loader, center, threshold, **options):
    model.eval()
    scores, predict, labels = [], [], []
    with torch.no_grad():
        for data in test_loader:
            inputs, target_name, target = data
            inputs = inputs.to(device)
            encoder_out = model(inputs)
            predict.append(encoder_out.cpu().numpy())
            scores.append(encoder_out.cpu().numpy())
            labels.extend(target)
    scores = np.concatenate(scores, 0)
    predict = np.concatenate(predict, 0)
    labels = np.array(labels)
    # np.savez('./log/PCfeature.npz', outputs_list=predict, label_list=labels)

    dist = get_dist(center, scores)
    labels_open = []
    for i in range(0, len(labels)):
        if labels[i] > len(options['train_typelist']):
            labels[i] = len(options['train_typelist'])
            labels_open.append(0)
        elif labels[i] == len(options['train_typelist']):
            labels_open.append(0)
        else:
            labels_open.append(1)
    labels_open = np.array(labels_open)
    predict = np.argmax(predict, axis=1)
    min_value_list = np.min(dist, axis=1)
    unknown_index = np.where(min_value_list > threshold)[0]
    labels_open_pred = predict
    for i in range(len(unknown_index)):
        labels_open_pred[unknown_index[i]] = len(options['train_typelist'])
    evaluate = classification_report(labels, labels_open_pred, output_dict=True)
    accuracy = evaluate['accuracy']
    f1_weighted = evaluate['weighted avg']['f1-score']
    f1_macro = evaluate['macro avg']['f1-score']
    AUROC_bin = roc_auc_score(labels_open, -np.min(dist, axis=1))
    # AUROC_bin = 0

    print(f"PC AUROC is %.3f" % (100*AUROC_bin))
    print(f"PC accuracy is %.3f" % (100*accuracy))
    print(f"PC Weighted F1 is %.3f" % (100*f1_weighted))
    print(f"PC Macro F1 is %.3f" % (100*f1_macro))
    print(f"_________________________________________")

    return 100*AUROC_bin,100*accuracy,100*f1_weighted,100*f1_macro


def get_dist(center, feature):
    features_square = np.sum(np.square(feature), axis=1)
    features_square = np.tile(features_square, (len(center[1]), 1)).T
    centers_square = np.sum(np.square(center), axis=0)
    centers_square = np.tile(centers_square, (len(feature), 1))
    features_into_centers = 2 * np.dot(feature, center)
    dist = features_square + centers_square - features_into_centers
    return dist


def get_center(model, device, data_loader, **options):
    label_list, outputs_list, scores = [], [], []
    feature_means = ()
    feature_cell = ()
    with torch.no_grad():
        for data in data_loader:
            inputs, _, target = data
            inputs = inputs.to(device)
            inputs = torch.squeeze(inputs, dim=1)
            encoder_out = model(inputs)
            encoder_out = encoder_out.cpu()
            outputs_list.append(encoder_out.cpu().numpy())
            label_list.extend(target)
    outputs_list = np.concatenate(outputs_list, 0)
    label_list = np.array(label_list)

    unique_label = set(label_list)
    class_num = len(unique_label)
    for label in unique_label:
        label_index = [index for index, element in enumerate(label_list) if element == label]
        feature = outputs_list[label_index, :]
        feature_cell += (feature,)
        feature_mean = np.mean(feature, axis=0)
        feature_means += (feature_mean,)
        features_square = np.sum(np.square(feature), axis=1)
        centers_square = np.sum(np.square(feature_mean))
        features_into_centers = 2 * np.dot(feature, feature_mean)
        dist = features_square + centers_square - features_into_centers
        scores = [scores, dist]
        scores = np.concatenate(scores, 0)
    sorted_values = np.sort(scores)
    threshold = sorted_values[int(0.97 * len(scores)-1)]
    threshold = np.array(threshold).squeeze()
    center = np.stack(feature_means, axis=1)
    print(threshold)
    return center, threshold


def main(options):
    seed_everything(options['seed'])
    device = torch.device('cuda:0')
    options['num_classes'] = len(options['train_typelist'])
    feat_dim = 256
    # Model
    print('==> Building model: '+options['model_name'])
    if options['dataset'] == 'radar':
        model = ResNet50(in_channels=2, num_classes=len(options['train_typelist']))
    elif options['dataset'] == 'ADSB':
        model = ResNet50(in_channels=3, num_classes=len(options['train_typelist']))

    options.update(
        {
            'use_gpu': True,
            'feat_dim': feat_dim,
        }
    )


    model = model.to(device)
    print(options['model_name'])
    backbone_weight_path = options['weight_path']
    model.load_state_dict(torch.load(backbone_weight_path))
    data_root_path = options['data_root'] + options['dataset']
    # Dataset
    if options['dataset'] == 'radar':
        valset = dataset_for_radar(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_radar(data_root_path, options['test_typelist'], mode_type='test')
    else:
        valset = dataset_for_ADSB(data_root_path, options['train_typelist'], mode_type='val')
        testset = dataset_for_ADSB(data_root_path, options['test_typelist'], mode_type='test')

    #  dataloader
    validloader = DataLoader(valset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All val data:", len(valset))
    testloader = DataLoader(testset, batch_size=options['batch_size'], shuffle=True, pin_memory=True)
    print("All test data:", len(testset))

    center,threshold = get_center(model, device, validloader, **options)
    AUROC_bin, accuracy, f1_weighted, f1_macro = test(model, device, testloader, center, threshold, **options)
    new_test_results = {
        'AUROC': [AUROC_bin],
        'Accuracy': [accuracy],
        'F1_weighted': [f1_weighted],
        'F1_macro': [f1_macro],
    }
    new_df = pd.DataFrame(new_test_results)
    output_file = 'TEST.xlsx'
    try:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            startrow = writer.sheets['Sheet1'].max_row
            new_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)
        print(f"New test results have been appended to {output_file}")
    except FileNotFoundError:
        new_df.to_excel(output_file, index=False)
        print(f"The file does not exist. A new file has been created: {output_file}")

    return 0

if __name__ == '__main__':
    args = parser.parse_args()
    options = vars(args)
    results = dict()
    options.update({
        'model_name': 'ResNet50'
    })
    from split import train_splits as train_splits
    from split import test_splits as test_splits
    test_list = ['radar', 'ADSB']
    for j in range(len(test_list)):
        torch.cuda.empty_cache()
        options.update(
            {
                'dataset': test_list[j]
            }
        )
        result_model_path = './log/PC/' + options['model_name'] + '/'+ options['dataset']
        files = os.listdir(result_model_path)

        for i in range(len(train_splits[options['dataset']])):
            train_split = train_splits[options['dataset']][i]
            test_split = test_splits[options['dataset']][i]
            weight_path = './log/PC/'+ options['model_name'] + '/' + options['dataset'] + '/' + files[i] + '/result_model/'+options['model_name']+'.pth'
            options.update(
                {
                    'train_typelist': train_split,  # [0,1,2,3,4,5,6]
                    'val_typelist': train_split,
                    'test_typelist': test_split,
                    'seed': 0,
                    'batch_size': 64,
                    'weight_path': weight_path,
                    'data_root': 'D:/wu/00data/',
                }
            )
            main(options)